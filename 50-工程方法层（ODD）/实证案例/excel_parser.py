
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional
from datetime import datetime
import re


class ExcelParser:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        
    def load(self, sheet_name: Optional[str] = None, sheet_index: int = 0):
        self.workbook = load_workbook(self.file_path, data_only=True)
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.worksheets[sheet_index]
        return self
    
    def parse_to_dict_list(self, header_row: int = 1, start_row: int = 2) -> List[Dict[str, Any]]:
        if not self.sheet:
            raise ValueError("Sheet not loaded. Call load() first.")
        
        headers = []
        for cell in self.sheet[header_row]:
            headers.append(cell.value)
        
        data = []
        for row in self.sheet.iter_rows(min_row=start_row, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = self._convert_type(value)
            data.append(row_dict)
        
        return data
    
    def parse_to_list(self, start_row: int = 1) -> List[List[Any]]:
        if not self.sheet:
            raise ValueError("Sheet not loaded. Call load() first.")
        
        data = []
        for row in self.sheet.iter_rows(min_row=start_row, values_only=True):
            data.append([self._convert_type(cell) for cell in row])
        
        return data
    
    def _convert_type(self, value: Any) -> Any:
        if value is None:
            return None
        
        if isinstance(value, (int, float, bool)):
            return value
        
        if isinstance(value, datetime):
            return value
        
        if isinstance(value, str):
            value = value.strip()
            
            if value == '':
                return None
            
            if value.lower() in ('true', 'false'):
                return value.lower() == 'true'
            
            if re.match(r'^-?\d+$', value):
                return int(value)
            
            if re.match(r'^-?\d+\.\d+$', value):
                return float(value)
            
            return value
        
        return value
    
    def get_sheet_names(self) -> List[str]:
        if not self.workbook:
            self.workbook = load_workbook(self.file_path, data_only=True)
        return self.workbook.sheetnames
    
    def close(self):
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


# 使用示例
if __name__ == "__main__":
    # 方式1: 使用上下文管理器
    with ExcelParser("data.xlsx") as parser:
        parser.load(sheet_index=0)
        data = parser.parse_to_dict_list()
        print(data)
    
    # 方式2: 手动管理
    parser = ExcelParser("data.xlsx")
    parser.load(sheet_name="Sheet1")
    data = parser.parse_to_list()
    parser.close()
