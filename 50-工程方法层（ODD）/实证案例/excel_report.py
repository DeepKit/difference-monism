"""
Excel报表生成器
支持数据写入、格式化、图表生成等功能
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict, Any, Optional, Union
import os
from datetime import datetime
import pandas as pd


class ExcelReportGenerator:
    """Excel报表生成器类"""
    
    def __init__(self, filename: str, overwrite: bool = False):
        """
        初始化报表生成器
        
        Args:
            filename: Excel文件名
            overwrite: 是否覆盖已存在的文件
        """
        self.filename = filename
        self.workbook = None
        self.current_sheet = None
        
        try:
            if os.path.exists(filename) and not overwrite:
                self.workbook = load_workbook(filename)
                self.current_sheet = self.workbook.active
            else:
                self.workbook = Workbook()
                self.current_sheet = self.workbook.active
        except Exception as e:
            raise Exception(f"初始化Excel文件失败: {str(e)}")
    
    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> bool:
        """
        创建新工作表
        
        Args:
            sheet_name: 工作表名称
            index: 插入位置
        
        Returns:
            是否创建成功
        """
        try:
            if sheet_name in self.workbook.sheetnames:
                self.current_sheet = self.workbook[sheet_name]
                return True
            
            if index is not None:
                self.current_sheet = self.workbook.create_sheet(sheet_name, index)
            else:
                self.current_sheet = self.workbook.create_sheet(sheet_name)
            return True
        except Exception as e:
            print(f"创建工作表失败: {str(e)}")
            return False
    
    def select_sheet(self, sheet_name: str) -> bool:
        """
        选择工作表
        
        Args:
            sheet_name: 工作表名称
        
        Returns:
            是否选择成功
        """
        try:
            if sheet_name in self.workbook.sheetnames:
                self.current_sheet = self.workbook[sheet_name]
                return True
            else:
                print(f"工作表 '{sheet_name}' 不存在")
                return False
        except Exception as e:
            print(f"选择工作表失败: {str(e)}")
            return False
    
    def write_data(self, data: Union[List[List], List[Dict], pd.DataFrame], 
                   start_row: int = 1, start_col: int = 1, 
                   include_header: bool = True) -> bool:
        """
        写入数据
        
        Args:
            data: 数据（列表、字典列表或DataFrame）
            start_row: 起始行
            start_col: 起始列
            include_header: 是否包含表头
        
        Returns:
            是否写入成功
        """
        try:
            if isinstance(data, pd.DataFrame):
                return self._write_dataframe(data, start_row, start_col, include_header)
            elif isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    return self._write_dict_list(data, start_row, start_col, include_header)
                else:
                    return self._write_list(data, start_row, start_col)
            else:
                print("不支持的数据格式")
                return False
        except Exception as e:
            print(f"写入数据失败: {str(e)}")
            return False
    
    def _write_list(self, data: List[List], start_row: int, start_col: int) -> bool:
        """写入二维列表数据"""
        try:
            for row_idx, row_data in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row_data, start=start_col):
                    self.current_sheet.cell(row=row_idx, column=col_idx, value=value)
            return True
        except Exception as e:
            print(f"写入列表数据失败: {str(e)}")
            return False
    
    def _write_dict_list(self, data: List[Dict], start_row: int, 
                        start_col: int, include_header: bool) -> bool:
        """写入字典列表数据"""
        try:
            if not data:
                return False
            
            headers = list(data[0].keys())
            current_row = start_row
            
            # 写入表头
            if include_header:
                for col_idx, header in enumerate(headers, start=start_col):
                    self.current_sheet.cell(row=current_row, column=col_idx, value=header)
                current_row += 1
            
            # 写入数据
            for row_data in data:
                for col_idx, header in enumerate(headers, start=start_col):
                    value = row_data.get(header, "")
                    self.current_sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            return True
        except Exception as e:
            print(f"写入字典数据失败: {str(e)}")
            return False
    
    def _write_dataframe(self, df: pd.DataFrame, start_row: int, 
                        start_col: int, include_header: bool) -> bool:
        """写入DataFrame数据"""
        try:
            current_row = start_row
            
            # 写入表头
            if include_header:
                for col_idx, header in enumerate(df.columns, start=start_col):
                    self.current_sheet.cell(row=current_row, column=col_idx, value=header)
                current_row += 1
            
            # 写入数据
            for _, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, start=start_col):
                    self.current_sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            return True
        except Exception as e:
            print(f"写入DataFrame失败: {str(e)}")
            return False
    
    def format_header(self, row: int = 1, start_col: int = 1, 
                     end_col: Optional[int] = None,
                     bg_color: str = "366092", font_color: str = "FFFFFF",
                     bold: bool = True, font_size: int = 11) -> bool:
        """
        格式化表头
        
        Args:
            row: 表头行号
            start_col: 起始列
            end_col: 结束列
            bg_color: 背景色
            font_color: 字体颜色
            bold: 是否加粗
            font_size: 字体大小
        
        Returns:
            是否格式化成功
        """
        try:
            if end_col is None:
                end_col = self.current_sheet.max_column
            
            for col in range(start_col, end_col + 1):
                cell = self.current_sheet.cell(row=row, column=col)
                cell.font = Font(bold=bold, color=font_color, size=font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            return True
        except Exception as e:
            print(f"格式化表头失败: {str(e)}")
            return False
    
    def format_cells(self, start_row: int, start_col: int, 
                    end_row: int, end_col: int,
                    font_size: int = 10, alignment: str = "left",
                    border: bool = True) -> bool:
        """
        格式化单元格区域
        
        Args:
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
            font_size: 字体大小
            alignment: 对齐方式
            border: 是否添加边框
        
        Returns:
            是否格式化成功
        """
        try:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = self.current_sheet.cell(row=row, column=col)
                    cell.font = Font(size=font_size)
                    cell.alignment = Alignment(horizontal=alignment, vertical="center")
                    if border:
                        cell.border = thin_border
            
            return True
        except Exception as e:
            print(f"格式化单元格失败: {str(e)}")
            return False
    
    def set_column_width(self, columns: Union[int, List[int]], 
                        width: Union[float, List[float]]) -> bool:
        """
        设置列宽
        
        Args:
            columns: 列号或列号列表
            width: 宽度或宽度列表
        
        Returns:
            是否设置成功
        """
        try:
            if isinstance(columns, int):
                columns = [columns]
                width = [width]
            elif isinstance(width, (int, float)):
                width = [width] * len(columns)
            
            for col, w in zip(columns, width):
                col_letter = get_column_letter(col)
                self.current_sheet.column_dimensions[col_letter].width = w
            
            return True
        except Exception as e:
            print(f"设置列宽失败: {str(e)}")
            return False
    
    def auto_fit_columns(self, start_col: int = 1, 
                        end_col: Optional[int] = None) -> bool:
        """
        自动调整列宽
        
        Args:
            start_col: 起始列
            end_col: 结束列
        
        Returns:
            是否调整成功
        """
        try:
            if end_col is None:
                end_col = self.current_sheet.max_column
            
            for col in range(start_col, end_col + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                
                for cell in self.current_sheet[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                self.current_sheet.column_dimensions[col_letter].width = adjusted_width
            
            return True
        except Exception as e:
            print(f"自动调整列宽失败: {str(e)}")
            return False
    
    def add_chart(self, chart_type: str, data_range: str, 
                 title: str = "", position: str = "E2",
                 categories_range: Optional[str] = None) -> bool:
        """
        添加图表
        
        Args:
            chart_type: 图表类型 (bar/line/pie)
            data_range: 数据范围
            title: 图表标题
            position: 图表位置
            categories_range: 分类范围
        
        Returns:
            是否添加成功
        """
        try:
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                print(f"不支持的图表类型: {chart_type}")
                return False
            
            chart.title = title
            data = Reference(self.current_sheet, range_string=data_range)
            
            if categories_range:
                cats = Reference(self.current_sheet, range_string=categories_range)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
            else:
                chart.add_data(data, titles_from_data=True)
            
            self.current_sheet.add_chart(chart, position)
            return True
        except Exception as e:
            print(f"添加图表失败: {str(e)}")
            return False
    
    def add_formula(self, row: int, col: int, formula: str) -> bool:
        """
        添加公式
        
        Args:
            row: 行号
            col: 列号
            formula: 公式（不含等号）
        
        Returns:
            是否添加成功
        """
        try:
            cell = self.current_sheet.cell(row=row, column=col)
            cell.value = f"={formula}"
            return True
        except Exception as e:
            print(f"添加公式失败: {str(e)}")
            return False
    
    def merge_cells(self, start_row: int, start_col: int, 
                   end_row: int, end_col: int) -> bool:
        """
        合并单元格
        
        Args:
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
        
        Returns:
            是否合并成功
        """
        try:
            start_cell = f"{get_column_letter(start_col)}{start_row}"
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            self.current_sheet.merge_cells(f"{start_cell}:{end_cell}")
            return True
        except Exception as e:
            print(f"合并单元格失败: {str(e)}")
            return False
    
    def add_data_validation(self, start_row: int, start_col: int,
                          end_row: int, end_col: int,
                          validation_type: str, formula: str) -> bool:
        """
        添加数据验证
        
        Args:
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
            validation_type: 验证类型 (list/whole/decimal/date/time/textLength/custom)
            formula: 验证公式
        
        Returns:
            是否添加成功
        """
        try:
            dv = DataValidation(type=validation_type, formula1=formula, allow_blank=True)
            self.current_sheet.add_data_validation(dv)
            
            start_cell = f"{get_column_letter(start_col)}{start_row}"
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            dv.add(f"{start_cell}:{end_cell}")
            
            return True
        except Exception as e:
            print(f"添加数据验证失败: {str(e)}")
            return False
    
    def freeze_panes(self, row: int = 2, col: int = 1) -> bool:
        """
        冻结窗格
        
        Args:
            row: 冻结行号
            col: 冻结列号
        
        Returns:
            是否冻结成功
        """
        try:
            cell = f"{get_column_letter(col)}{row}"
            self.current_sheet.freeze_panes = cell
            return True
        except Exception as e:
            print(f"冻结窗格失败: {str(e)}")
            return False
    
    def save(self, filename: Optional[str] = None) -> bool:
        """
        保存Excel文件
        
        Args:
            filename: 文件名（可选）
        
        Returns:
            是否保存成功
        """
        try:
            save_path = filename if filename else self.filename
            self.workbook.save(save_path)
            print(f"文件已保存: {save_path}")
            return True
        except Exception as e:
            print(f"保存文件失败: {str(e)}")
            return False
    
    def close(self):
        """关闭工作簿"""
        try:
            if self.workbook:
                self.workbook.close()
        except Exception as e:
            print(f"关闭工作簿失败: {str(e)}")


# 使用示例
if __name__ == "__main__":
    # 创建报表生成器
    report = ExcelReportGenerator("销售报表.xlsx", overwrite=True)
    
    # 示例数据
    sales_data = [
        {"日期": "2024-01-01", "产品": "产品A", "销量": 100, "金额": 5000},
        {"日期": "2024-01-02", "产品": "产品B", "销量": 150, "金额": 7500},
        {"日期": "2024-01-03", "产品": "产品C", "销量": 200, "金额": 10000},
        {"日期": "2024-01-04", "产品": "产品A", "销量": 120, "金额": 6000},
        {"日期": "2024-01-05", "产品": "产品B", "销量": 180, "金额": 9000},
    ]
    
    # 写入数据
    report.write_data(sales_data, start_row=1, start_col=1, include_header=True)
    
    # 格式化表头
    report.format_header(row=1, start_col=1, end_col=4)
    
    # 格式化数据区域
    report.format_cells(2, 1, 6, 4, alignment="center", border=True)
    
    # 自动调整列宽
    report.auto_fit_columns()
    
    # 添加合计行
    report.current_sheet.cell(row=7, column=1, value="合计")
    report.add_formula(7, 3, "SUM(C2:C6)")
    report.add_formula(7, 4, "SUM(D2:D6)")
    
    # 冻结首行
    report.freeze_panes(row=2, col=1)
    
    # 保存文件
    report.save()
    report.close()
    
    print("报表生成完成！")