from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any


class ExcelReportGenerator:
    def __init__(self, filename: str = "report.xlsx"):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def set_sheet_name(self, name: str):
        """设置当前工作表名称"""
        self.ws.title = name
        
    def add_header(self, headers: List[str], row: int = 1):
        """添加表头"""
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    def add_data(self, data: List[List[Any]], start_row: int = 2):
        """添加数据行"""
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                self.ws.cell(row=row_idx, column=col_idx, value=value)
                
    def add_dict_data(self, data: List[Dict[str, Any]], headers: List[str] = None, start_row: int = 1):
        """添加字典格式数据（自动生成表头）"""
        if not data:
            return
            
        if headers is None:
            headers = list(data[0].keys())
            
        self.add_header(headers, start_row)
        
        rows = [[row.get(key, "") for key in headers] for row in data]
        self.add_data(rows, start_row + 1)
        
    def auto_adjust_width(self):
        """自动调整列宽"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
            
    def save(self):
        """保存文件"""
        self.wb.save(self.filename)
        
    def create_new_sheet(self, name: str):
        """创建新工作表"""
        self.ws = self.wb.create_sheet(title=name)
        return self.ws


# 使用示例
if __name__ == "__main__":
    # 创建报表生成器
    report = ExcelReportGenerator("销售报表.xlsx")
    report.set_sheet_name("月度销售")
    
    # 方式1：使用列表数据
    headers = ["产品", "销量", "金额"]
    data = [
        ["产品A", 100, 5000],
        ["产品B", 200, 8000],
        ["产品C", 150, 6000]
    ]
    report.add_header(headers)
    report.add_data(data)
    
    # 方式2：使用字典数据（新工作表）
    report.create_new_sheet("客户信息")
    customers = [
        {"姓名": "张三", "电话": "13800138000", "金额": 5000},
        {"姓名": "李四", "电话": "13900139000", "金额": 8000}
    ]
    report.add_dict_data(customers)
    
    report.auto_adjust_width()
    report.save()