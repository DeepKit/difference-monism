
"""
Excel文件解析工具
支持读取.xlsx格式的Excel文件
"""

from typing import List, Dict, Any, Optional, Union
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self, file_path: str):
        """
        初始化Excel解析器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.active_sheet = self.workbook.active
    
    def get_sheet_names(self) -> List[str]:
        """获取所有工作表名称"""
        return self.workbook.sheetnames
    
    def select_sheet(self, sheet_name: str) -> None:
        """
        选择工作表
        
        Args:
            sheet_name: 工作表名称
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        self.active_sheet = self.workbook[sheet_name]
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """
        获取单元格值
        
        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）
        """
        return self.active_sheet.cell(row, col).value
    
    def get_row(self, row_num: int, start_col: int = 1, end_col: Optional[int] = None) -> List[Any]:
        """
        获取指定行的数据
        
        Args:
            row_num: 行号
            start_col: 起始列号
            end_col: 结束列号（None表示到最后一列）
        """
        if end_col is None:
            end_col = self.active_sheet.max_column
        
        return [self.active_sheet.cell(row_num, col).value 
                for col in range(start_col, end_col + 1)]
    
    def get_column(self, col_num: int, start_row: int = 1, end_row: Optional[int] = None) -> List[Any]:
        """
        获取指定列的数据
        
        Args:
            col_num: 列号
            start_row: 起始行号
            end_row: 结束行号（None表示到最后一行）
        """
        if end_row is None:
            end_row = self.active_sheet.max_row
        
        return [self.active_sheet.cell(row, col_num).value 
                for row in range(start_row, end_row + 1)]
    
    def get_all_data(self, has_header: bool = True) -> List[List[Any]]:
        """
        获取所有数据
        
        Args:
            has_header: 是否包含表头
        """
        data = []
        start_row = 1 if has_header else 2
        
        for row in self.active_sheet.iter_rows(min_row=start_row, 
                                               values_only=True):
            data.append(list(row))
        
        return data
    
    def get_data_as_dict(self, header_row: int = 1) -> List[Dict[str, Any]]:
        """
        将数据转换为字典列表（以表头为键）
        
        Args:
            header_row: 表头所在行号
        """
        headers = self.get_row(header_row)
        data = []
        
        for row_num in range(header_row + 1, self.active_sheet.max_row + 1):
            row_data = self.get_row(row_num)
            row_dict = {headers[i]: row_data[i] for i in range(len(headers))}
            data.append(row_dict)
        
        return data
    
    def get_range(self, start_row: int, start_col: int, 
                  end_row: int, end_col: int) -> List[List[Any]]:
        """
        获取指定范围的数据
        
        Args:
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
        """
        data = []
        for row in range(start_row, end_row + 1):
            row_data = [self.active_sheet.cell(row, col).value 
                       for col in range(start_col, end_col + 1)]
            data.append(row_data)
        
        return data
    
    def search_value(self, value: Any, sheet_name: Optional[str] = None) -> List[tuple]:
        """
        搜索指定值的位置
        
        Args:
            value: 要搜索的值
            sheet_name: 工作表名称（None表示当前工作表）
        
        Returns:
            包含(行号, 列号)的列表
        """
        if sheet_name:
            self.select_sheet(sheet_name)
        
        positions = []
        for row in range(1, self.active_sheet.max_row + 1):
            for col in range(1, self.active_sheet.max_column + 1):
                if self.active_sheet.cell(row, col).value == value:
                    positions.append((row, col))
        
        return positions
    
    def get_sheet_info(self) -> Dict[str, Any]:
        """获取当前工作表信息"""
        return {
            'sheet_name': self.active_sheet.title,
            'max_row': self.active_sheet.max_row,
            'max_column': self.active_sheet.max_column,
            'dimensions': self.active_sheet.dimensions
        }
    
    def close(self) -> None:
        """关闭工作簿"""
        self.workbook.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器退出"""
        self.close()


# 使用示例
if __name__ == "__main__":
    # 基本使用
    with ExcelParser("data.xlsx") as parser:
        # 获取所有工作表名称
        sheets = parser.get_sheet_names()
        print(f"工作表: {sheets}")
        
        # 获取工作表信息
        info = parser.get_sheet_info()
        print(f"工作表信息: {info}")
        
        # 读取单元格
        value = parser.get_cell_value(1, 1)
        print(f"A1单元格: {value}")
        
        # 读取整行
        row_data = parser.get_row(1)
        print(f"第1行: {row_data}")
        
        # 读取整列
        col_data = parser.get_column(1)
        print(f"第1列: {col_data}")
        
        # 读取所有数据
        all_data = parser.get_all_data()
        print(f"所有数据: {all_data[:5]}")  # 只显示前5行
        
        # 转换为字典格式
        dict_data = parser.get_data_as_dict()
        print(f"字典格式: {dict_data[:3]}")  # 只显示前3条
        
        # 搜索值
        positions = parser.search_value("目标值")
        print(f"找到位置: {positions}")
